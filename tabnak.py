import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, Alignment

# آدرس سایت
url = "https://www.tabnak.ir"
headers = {"User-Agent": "Mozilla/5.0"}

response = requests.get(url, headers=headers)

if response.status_code == 200:
    soup = BeautifulSoup(response.text, "html.parser")
    titles = soup.find_all("a", class_="title5")

    print(f"✅ تعداد عناوین پیدا شده: {len(titles)}")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "عناوین سایت تابناک"

    # تنظیم عرض ستون‌ها
    ws.column_dimensions['A'].width = 100  # عنوان
    ws.column_dimensions['B'].width = 20   # لینک

    bold_font = Font(bold=True, name='Tahoma', size=12)
    link_font = Font(color="0000FF", underline="single")
    right_align = Alignment(horizontal='right')

    for i, tag in enumerate(titles, 1):
        title_text = tag.get_text(strip=True) + " (تابناک)"
        href = tag.get("href", "")

        if href.startswith("/"):
            href = "https://www.tabnak.ir" + href

        # نوشتن عنوان
        cell_title = ws.cell(row=i, column=1, value=title_text)
        cell_title.font = bold_font
        cell_title.alignment = right_align

        # نوشتن لینک قابل کلیک
        if href:
            cell_link = ws.cell(row=i, column=2, value="مشاهده")
            cell_link.hyperlink = href
            cell_link.font = link_font
            cell_link.alignment = right_align

        ws.row_dimensions[i].height = 25

    wb.save("tabnak_titles_with_links.xlsx")
    print("📁 فایل tabnak_titles_with_links.xlsx با موفقیت ذخیره شد.")
else:
    print("❌ خطا در دریافت اطلاعات از سایت")






   # آدرس سایت
url = "https://www.isna.ir/archive"
headers = {"User-Agent": "Mozilla/5.0"}

response = requests.get(url, headers=headers)

if response.status_code == 200:
    soup = BeautifulSoup(response.text, "html.parser")
    titles = soup.find_all("a", class_="desc")

    print(f"✅ تعداد عناوین پیدا شده: {len(titles)}")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "آخرین اخبار ایسنا"

    # تنظیم عرض ستون‌ها
    ws.column_dimensions['A'].width = 100  # عنوان
    ws.column_dimensions['B'].width = 20   # لینک

    bold_font = Font(bold=True, name='Tahoma', size=12)
    link_font = Font(color="0000FF", underline="single")
    right_align = Alignment(horizontal='right')

    for i, tag in enumerate(titles, 1):
        title_text = tag.get_text(strip=True) + " (ایسنا)"
        href = tag.get("href", "")

        if href.startswith("/"):
            href = "https://www.isna.ir/archive" + href

        # نوشتن عنوان
        cell_title = ws.cell(row=i, column=4, value=title_text)
        cell_title.font = bold_font
        cell_title.alignment = right_align

        # نوشتن لینک قابل کلیک
        if href:
            cell_link = ws.cell(row=i, column=5, value="مشاهده")
            cell_link.hyperlink = href
            cell_link.font = link_font
            cell_link.alignment = right_align

        ws.row_dimensions[i].height = 25

    wb.save("news_titles_with_links.xlsx")
    print("📁 فایل news_titles_with_links.xlsx با موفقیت ذخیره شد.")
else:
    print("❌ خطا در دریافت اطلاعات از سایت") 