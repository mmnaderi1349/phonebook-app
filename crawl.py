import requests
from bs4 import BeautifulSoup
import pandas as pd
from datetime import datetime, timedelta
import pytz
import time
import os
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

# دیکشنری دسته‌ها با فرا‌دسته‌ها
super_categories = {
    "لوازم آرایشی": {
        "https://liateam.ir/categories/148/sub/149": "🔽 آرایش چشم",
    },
    "مراقبت پوست": {
        "https://liateam.ir/categories/90/sub/126": "🔽 شوینده ها و پاک کننده ها",
        "https://liateam.ir/categories/90/sub/112": "🔽 صابون",
        "https://liateam.ir/categories/90/sub/113": "🔽 پن",
        "https://liateam.ir/categories/90/sub/114": "🔽 کرم",
        "https://liateam.ir/categories/10/sub/114?page=2": "🔽 کرم صفحه دو",
        "https://liateam.ir/categories/90/sub/146": "🔽 بالم لب",
        "https://liateam.ir/categories/90/sub/131": "🔽 ضد آفتاب",
        "https://liateam.ir/categories/90/sub/128": "🔽 ماسک صورت",
        "https://liateam.ir/categories/90/sub/116": "🔽 ناحیه چشم",
        "https://liateam.ir/categories/90/sub/115": "🔽 سرم و روغن",
        "https://liateam.ir/categories/10/sub/115?page=2": "🔽 سرم و روغن صفحه دو ",
        "https://liateam.ir/categories/90/sub/139": "🔽 کیت"
    },
    "مراقبت مو و ناخن": {
        "https://liateam.ir/categories/90/sub/100": "🔽 شامپو",
        "https://liateam.ir/categories/90/sub/101": "🔽 ماسک مو",
        "https://liateam.ir/categories/90/sub/102": "🔽 تونیک",
        "https://liateam.ir/categories/90/sub/104": "🔽 سرم مو",
        "https://liateam.ir/categories/90/sub/103": "🔽 روغن مو",
        "https://liateam.ir/categories/90/sub/122": "🔽 کیت رویش مجدد موی سر",
        "https://liateam.ir/categories/90/sub/147": "🔽 مراقبت از ناخن"
    },
    "مراقبت بدن": {
        "https://liateam.ir/categories/90/sub/117": "🔽 شامپوی بدن",
        "https://liateam.ir/categories/90/sub/119": "🔽 کرم",
        "https://liateam.ir/categories/90/sub/118": "🔽 روغن و لوسیون",
        "https://liateam.ir/categories/90/sub/142": "🔽 خوشبو کننده و ضد تعریق",
        "https://liateam.ir/categories/90/sub/135": "🔽 مایع دستشویی"
    },
    "عطر و آرایشی": {
        "https://liateam.ir/categories/91/sub/134": "🔽 کرم DD",
        "https://liateam.ir/categories/91/sub/132": "🔽 کرم BB",
        "https://liateam.ir/categories/91/sub/133": "🔽 کرم CC",
        "https://liateam.ir/categories/91/sub/127": "🔽 کانسیلر",
        "https://liateam.ir/categories/91/sub/137": "🔽 پرفیوم و ادوپرفیوم",
        "https://liateam.ir/categories/91/sub/138": "🔽 بادی میست",
        "https://liateam.ir/categories/91/sub/138?page=2": "🔽 بادی میست صفحه دو",
        "https://liateam.ir/categories/91/sub/144": "🔽 خوشبوکننده هوا"
    },
    "مراقبت دهان و دندان": {
        "https://liateam.ir/categories/140/sub/143": "🔽 خمیر دندان",
        "https://liateam.ir/categories/140/sub/141": "🔽 دهانشویه",
        "https://liateam.ir/categories/140/sub/145": "🔽 خوشبو کننده دهان"
    }
}

def fetch_data(url):
    response = requests.get(url)
    if response.status_code == 200:
        soup = BeautifulSoup(response.text, 'html.parser')
        titles = [item.get_text(strip=True) for item in soup.find_all('span', class_='chakra-text css-2kfigy')]
        prices = [item.get_text(strip=True) for item in soup.find_all('div', class_='css-mqrh1k')]
        if len(titles) != len(prices):
            print(f"تعداد عنوان‌ها و قیمت‌ها برابر نیست: {url}")
            return None
        iran_time = datetime.now(pytz.timezone('Asia/Tehran')).strftime('%Y-%m-%d %H:%M:%S')
        return pd.DataFrame({'محصول': titles, 'قیمت': prices, 'زمان ثبت': iran_time})
    else:
        print(f"❌ خطا در دریافت داده از: {url} - Status code: {response.status_code}")
        return None

with pd.ExcelWriter('output_super_categories.xlsx', engine='openpyxl') as writer:
    startrow = 0
    for super_cat, links in super_categories.items():
        # درج فرا‌دسته
        super_cat_row = pd.DataFrame([{'محصول': super_cat, 'قیمت': '', 'زمان ثبت': ''}])
        super_cat_row.to_excel(writer, index=False, header=False, sheet_name='Sheet1', startrow=startrow)
        startrow += 1

        for url, sub_cat in links.items():
            df = fetch_data(url)
            if df is not None:
                cat_row = pd.DataFrame([{'محصول': sub_cat, 'قیمت': '', 'زمان ثبت': ''}])
                cat_row.to_excel(writer, index=False, header=False, sheet_name='Sheet1', startrow=startrow)
                startrow += 1
                df.to_excel(writer, index=False, sheet_name='Sheet1', startrow=startrow)
                startrow += len(df) + 2
                time.sleep(20)

    # قالب‌بندی اکسل
    workbook = writer.book
    sheet = writer.sheets['Sheet1']

    column_widths = [70, 20, 25]
    for i, width in enumerate(column_widths, start=1):
        sheet.column_dimensions[get_column_letter(i)].width = width

    for row in sheet.iter_rows():
        for cell in row:
            if cell.value:
                val = str(cell.value)
                if val in super_categories:  # سطر فرا‌دسته
                    cell.font = Font(bold=True, size=14, color="008000")
                    cell.alignment = Alignment(horizontal='center')
                elif val.startswith("🔽"):  # سطر دسته
                    cell.font = Font(bold=True, color="FF0000")
                    cell.alignment = Alignment(horizontal='center')
                elif cell.column == 2:
                    cell.alignment = Alignment(horizontal='center')

print("✅ فایل اکسل با فرا‌دسته‌ها و قالب‌بندی کامل ساخته شد.")