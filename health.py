import requests
from bs4 import BeautifulSoup
import openpyxl

# مرحله ۱: دریافت محتوای HTML از سایت با مدیریت خطا
url = "https://www.health.com/trending-8782982"
headers = {
    "User-Agent": "Mozilla/5.0"
}

try:
    response = requests.get(url, headers=headers, timeout=10)
    response.raise_for_status()
except requests.exceptions.HTTPError as http_err:
    print(f"❌ خطای HTTP هنگام اتصال به سایت: {http_err}")
    exit()
except requests.exceptions.ConnectionError:
    print("❌ خطای اتصال: به نظر می‌رسد به سایت مقصد وصل نمی‌شوید (شاید فیلتر یا اینترنت مشکل دارد).")
    exit()
except requests.exceptions.Timeout:
    print("❌ زمان اتصال به پایان رسید. سایت یا اینترنت کند است.")
    exit()
except requests.exceptions.RequestException as err:
    print(f"❌ خطای غیرمنتظره هنگام اتصال: {err}")
    exit()

# مرحله ۲: پردازش محتوای HTML
soup = BeautifulSoup(response.content, "html.parser")

# مرحله ۳: پیدا کردن عناوین
titles = soup.find_all(class_="card__title-text")

# مرحله ۴: پیدا کردن تصاویر
image_divs = soup.find_all("div", class_="img-placeholder")

# مرحله ۴.۵: پیدا کردن پیوندها
link_cards = soup.find_all("a", class_="comp mntl-card-list-items mntl-universal-card mntl-document-card mntl-card card card--no-image")
hrefs = []

for a in link_cards:
    if a.has_attr("href"):
        href = a["href"]
        if href.startswith("/"):
            href = "https://www.health.com" + href  # تکمیل لینک‌های ناقص
        hrefs.append(href)
    else:
        hrefs.append("")

# مرحله ۵: ساخت فایل اکسل
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Health Titles"
ws.append(["عنوان", "آدرس تصویر", "پیوند"])

# مرحله ۶: پر کردن اطلاعات
for i in range(len(titles)):
    title_text = titles[i].get_text(strip=True)
    image_url = ""

    if i < len(image_divs):
        img_tag = image_divs[i].find("img")
        if img_tag and img_tag.has_attr("src"):
            image_url = img_tag["src"]

    href = hrefs[i] if i < len(hrefs) else ""

    # درج لینک واقعی قابل کلیک
    if href:
        cell = openpyxl.cell.cell.Cell(ws, value=None)
        ws.append([title_text, image_url, None])
        ws.cell(row=ws.max_row, column=3).value = f'=HYPERLINK("{href}", "مشاهده")'
    else:
        ws.append([title_text, image_url, ""])

# مرحله ۷: ذخیره فایل
wb.save("health_titles.xlsx")
print("✅ فایل 'health_titles.xlsx' با لینک‌های قابل کلیک ساخته شد.")